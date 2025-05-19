import asyncio
import nest_asyncio
import websockets
import json
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, ContextTypes
from telegram.ext.filters import TEXT
from openpyxl import Workbook, load_workbook
import os

nest_asyncio.apply()
WEBSOCKET_URL = "wss://ws.binaryws.com/websockets/v3?app_id=62714"
TELEGRAM_BOT_TOKEN = '7833833692:AAEfFSogXi1d0i3up9rrMMqrWpLWRpV8Cbc'
TRADING_KEY = "mizzf183B9B3733C58BdAC15Df"

LOG_FILE = "user_activitybasic.xlsx"

if not os.path.exists(LOG_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "UserActivity"
    ws.append(["User", "Initial Balance", "Loss/Profit", "Total Profit", "Final Balance"])
    wb.save(LOG_FILE)
    
def log_user_activity(user_name, initial_balance, profit_loss, total_profit, final_balance):
    wb = load_workbook(LOG_FILE)
    ws = wb["UserActivity"]
    ws.append([user_name, initial_balance, profit_loss, total_profit, final_balance])
    wb.save(LOG_FILE)

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    disclaimer_text = (
        "⚠️ Disclaimer: Trading Involves Risk\n\n"
        "Trading binary options and other financial instruments on platforms like Deriv involves significant risk of loss and may not be suitable for all investors.\n\n"
        "📉 You could lose all your invested capital.\n"
        "📊 Ensure you fully understand the risks before proceeding.\n\n"
        "*Do you agree to the terms and wish to continue using the Autonix bot?*"
    )

    keyboard = [
        [InlineKeyboardButton("✅ I Agree", callback_data="agree_terms")],
        [InlineKeyboardButton("❌ I Disagree", callback_data="disagree_terms")],
        [InlineKeyboardButton("📞 Contact Customer Care", url="https://t.me/autonix_assistant_bot")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(disclaimer_text, parse_mode="Markdown", reply_markup=reply_markup)


async def handle_terms_agreement(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    user_data = context.user_data

    if query.data == "agree_terms":
        user_data["agreed_to_terms"] = True
        await query.edit_message_text(
           "✅ You have agreed to the trading terms.\n\n"
           "\U0001F680 Welcome to Autonix' Lite Trading Bot! \U0001F680\n\n"
           "We're thrilled to have you onboard.\n\n"
           "Your trading journey is about to be *revolutionized to the next level beyond your imaginations!* \U0001F4C8\n"
           "\U0001F4B0 To get started:\n"
           "✅ Ensure you have a *Deriv API token\n"
           "✅ Use your *exclusive trading key\n"
           "✅ Start trading & *maximize your profits!\n\n"
           "✅ Configure your stake amount, take profit, and stop loss\n"
           "✅ Select a market and initiate trading with precision\n\n"
            "💼 Designed for traders who value automation, accuracy, and control.\n\n"
            "🔒 Your data is private. Your trades are yours.\n\n"
            "🧠 Powered by Autonix — Smart Trading Simplified."
            "\U0001F91D Best of luck, and happy trading! \U0001F3C6\n\n"
            "You may now use the bot. Use /trade to get started."
        )
    else:
        user_data["agreed_to_terms"] = False
        await query.edit_message_text(
            "❌ You must agree to the disclaimer to use this bot.\n"
            "Restart anytime with /start."
        )



async def trade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.user_data.get("agreed_to_terms"):
        await update.message.reply_text(
            "⚠️ You must agree to the disclaimer before using this feature.\n"
            "Please use /start and accept the terms to proceed."
        )
        return

    keyboard = [
        [
            InlineKeyboardButton("📝 Create Deriv Account", url="https://track.deriv.com/_AmZyUaWfEMJB4VdSfJsOp2Nd7ZgqdRLk/1"),
            InlineKeyboardButton("🔐 Get API Token", url="https://app.deriv.com/account/api-token")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🔑 Please enter your Deriv API token to proceed.\n\n"
        "If you don't have one, use the buttons below to create an account or generate a token.",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    context.user_data["next_step"] = "api_token"

async def show_commands(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    commands_text = (
        "📜 Available Commands:\n\n"
        "🔹 /start - Introduction to the bot\n"
        "🔹 /trade - Begin trading authentication\n"
        "🔹 /commands - Show this list of commands\n"
    )
    await update.message.reply_text(commands_text, parse_mode="Markdown")

async def handle_user_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_input = update.message.text
    next_step = context.user_data.get("next_step")
    
    if not next_step:
        await update.message.reply_text("Please start the bot using /start and then click /trade.")
        return
    
    if next_step == "api_token":
        api_token = user_input
        context.user_data["api_token"] = api_token
        async with websockets.connect(WEBSOCKET_URL) as websocket:
            await websocket.send(json.dumps({"authorize": api_token}))
            response = await websocket.recv()
            auth_data = json.loads(response)
            
            if "error" in auth_data:
                await update.message.reply_text(f"Authorization failed: {auth_data['error']['message']}")
                return
            
            user_name = auth_data.get("authorize", {}).get("fullname", "User")
            account_balance = auth_data.get("authorize", {}).get("balance", 0)
            context.user_data["user_name"] = user_name
            context.user_data["account_balance"] = account_balance
            
            await update.message.reply_text(
                f"✅ API Token Authenticated Successfully! 🎉\n\n"
                f"👤 User: {user_name}\n"
                f"💰 Balance: {account_balance} USD\n\n"
                "🔑 Now, enter your trading key to continue.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("I don't have a trading key", url="https://t.me/autonix_lite_key_bot")]])
            )
        context.user_data["next_step"] = "trading_key"


    elif next_step == "trading_key":
        if user_input == TRADING_KEY:
            await update.message.reply_text(
                "✅ Trading key authenticated successfully! 🎉\n\n"
                "📊 Select the trading symbol you want to trade:",
                reply_markup=symbol_selection_keyboard()
            )
            context.user_data["next_step"] = "symbol"
        else:
            await update.message.reply_text("Invalid key. Please enter a correct trading key.")



    elif next_step == "stake":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the stake amount.")
            return
        context.user_data["stake"] = float(user_input)
        await update.message.reply_text("Stake saved! Now, enter your take 💰profit target✅.")
        context.user_data["next_step"] = "take_profit"

    elif next_step == "take_profit":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the take profit target.")
            return
        context.user_data["take_profit"] = float(user_input)
        await update.message.reply_text("Take profit saved! Now, enter your stop loss target.")
        context.user_data["next_step"] = "stop_loss"

    elif next_step == "stop_loss":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the stop loss target.")
            return
        context.user_data["stop_loss"] = float(user_input)
        await update.message.reply_text(
            "Setup complete! Use the buttons below to start or stop trading.",
            reply_markup=start_stop_trading_keyboard(),
        )
        context.user_data["next_step"] = None

def symbol_selection_keyboard():
    buttons = [
        [InlineKeyboardButton("volatility 10 (1s)", callback_data="1HZ10V"), InlineKeyboardButton("volatility 10", callback_data="R_10")],
        [InlineKeyboardButton("volatility 25 (1s)", callback_data="1HZ25V"), InlineKeyboardButton("volatility 25", callback_data="R_25")],
        [InlineKeyboardButton("volatility 50 (1s)", callback_data="1HZ50V"), InlineKeyboardButton("volatility 50", callback_data="R_50")],
        [InlineKeyboardButton("volatility 75 (1s)", callback_data="1HZ75V"), InlineKeyboardButton("volatility 75", callback_data="R_75")],
        [InlineKeyboardButton("volatility 100 (1s)", callback_data="1HZ100V"), InlineKeyboardButton("volatility 100", callback_data="R_100")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_symbol_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    symbol = query.data

    context.user_data["symbol"] = symbol
    await query.edit_message_text(
        text=f"Symbol selected: {symbol}\n\n"
        "Please enter your stake amount (e.g., 10)."
    )
    context.user_data["next_step"] = "stake"

def start_stop_trading_keyboard():
    buttons = [
        [InlineKeyboardButton("🚀 Start Trading", callback_data="start_trading")],
        [InlineKeyboardButton("🛑 Stop Trading", callback_data="stop_trading")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_start_trading(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    user_info = context.user_data
    api_token = user_info.get("api_token")
    symbol = user_info.get("symbol")
    initial_stake = user_info.get("stake")
    take_profit = user_info.get("take_profit")
    stop_loss = user_info.get("stop_loss")
    user_name = user_info.get("user_name", "User")

    if not all([api_token, symbol, initial_stake, take_profit, stop_loss]):
        await query.edit_message_text("Error: Missing trading setup information.")
        return

    balance = 0
    stake = initial_stake
    trading_active = True

    async with websockets.connect(WEBSOCKET_URL) as websocket:
        # Authorization
        await websocket.send(json.dumps({"authorize": api_token}))
        auth_response = await websocket.recv()
        auth_data = json.loads(auth_response)

        if "error" in auth_data:
            await context.bot.send_message(
                chat_id=query.message.chat_id,
                text=f"Authorization failed: {auth_data['error']['message']}"
            )
            return

        initial_balance = auth_data.get("authorize", {}).get("balance", 0)
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=f"✅ Authorized successfully! Initial Balance: {initial_balance} USD\nTrading has started."
        )

        while trading_active and -stop_loss < balance < take_profit:
            contract_type = "DIGITUNDER" if stake % 2 == 0 else "DIGITOVER"
            digit = 8 if contract_type == "DIGITUNDER" else 1

            # Propose Trade
            proposal_request = {
                "proposal": 1,
                "amount": round(stake, 2),
                "basis": "stake",
                "contract_type": contract_type,
                "currency": "USD",
                "duration": 1,
                "duration_unit": "t",
                "symbol": symbol,
                "barrier": str(digit),
            }
            await websocket.send(json.dumps(proposal_request))
            response = await websocket.recv()
            proposal_response = json.loads(response)

            if "proposal" in proposal_response:
                contract_id = proposal_response["proposal"]["id"]

                # Buy Contract
                await websocket.send(json.dumps({"buy": contract_id, "price": round(stake, 2)}))
                purchase_response = await websocket.recv()
                purchase_data = json.loads(purchase_response)

                if "error" in purchase_data:
                    await query.edit_message_text(f"Purchase error: {purchase_data['error']['message']}")
                    break

                trade_id = purchase_data.get("buy", {}).get("contract_id")
                if trade_id:
                    while True:
                        # Monitor Contract
                        await websocket.send(json.dumps({"proposal_open_contract": 1, "contract_id": trade_id}))
                        result_response = await websocket.recv()
                        result_data = json.loads(result_response)

                        if "proposal_open_contract" in result_data:
                            contract = result_data["proposal_open_contract"]
                            if contract["is_sold"]:
                                trade_result = contract["profit"]
                                balance += trade_result
                                account_balance = initial_balance + balance

                                # Log User Activity
                                profit_loss = trade_result
                                total_profit = balance
                                final_balance = account_balance
                                log_user_activity(user_name, initial_balance, profit_loss, total_profit, final_balance)

                                # Handle Trade Outcome
                                if trade_result > 0:
                                    await context.bot.send_message(
                                        chat_id=query.message.chat_id,
                                        text=(
                                            f"🎉 Hurray, {user_name}!!! You Won: Profit = {trade_result} USD\n"
                                            f"Total Profit = {balance} USD\n"
                                            f"Account Balance = {account_balance} USD"
                                        )
                                    )
                                    stake = initial_stake
                                else:
                                    await context.bot.send_message(
                                        chat_id=query.message.chat_id,
                                        text=(
                                            f"😞 Bad luck, {user_name}!!! You Lost: Loss = {abs(trade_result)} USD\n"
                                            f"Total Loss = {abs(trade_result)} USD\n"
                                            f"Account Balance = {account_balance} USD"
                                        )
                                    )
                                    stake = max(stake * 4.3, 0.35)
                                break

            await asyncio.sleep(1)

    # Handle End of Trading Session
    if balance >= take_profit:
        keyboard = [
            [InlineKeyboardButton("🌟 Get all our tools", url="https://t.me/autonix_bot")],
            [InlineKeyboardButton("🌟 Trade using our Pro Bot", url="https://t.me/autonix_pro_bot")],
            [InlineKeyboardButton("🌟 Join Our Telegram Channel", url="https://t.me/autonix001")],
            [InlineKeyboardButton("🌐 Visit Our Trading Website", url="https://twbbestbots.com")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=(
                f"<b>🎉 Take Profit Reached!</b>\n\n"
                f"<i>Congratulations, {user_name}!</i>\n"
                f"Your trading session has <u>successfully ended</u>.\n\n"
                f"<b>🚀 Next Steps:</b>\n"
                f"• Check out our Telegram channel for updates.\n"
                f"• Explore our trading website for strategies.\n"
            ),
            parse_mode="HTML",
            reply_markup=reply_markup
        )

    elif balance <= -stop_loss:
        keyboard = [
            [InlineKeyboardButton("🌟 Join Our Telegram Channel", url="https://t.me/twbbestbot")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=(
                f"<b>🚨 Stop Loss Reached!</b>\n\n"
                f"<i>Dear {user_name},</i>\n"
                f"Unfortunately, your trading session has <u>ended</u> due to hitting the stop loss.\n\n"
                f"<b>📈 Tips:</b>\n"
                f"• Join our Telegram channel for guidance.\n"
                f"• Visit our website to refine your strategy.\n"
            ),
            parse_mode="HTML",
            reply_markup=reply_markup
        )

   




async def handle_stop_trading(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    user_name = context.user_data.get("user_name", "User")
    await query.edit_message_text(f"Trading stopped for {user_name}.")

def start_basic_bot():
    print("Starting the lite bot...")
    application = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(handle_terms_agreement, pattern="^(agree_terms|disagree_terms)$"))
    application.add_handler(CommandHandler("trade", trade))
    application.add_handler(CommandHandler("commands", show_commands))
    application.add_handler(MessageHandler(TEXT, handle_user_input))
    application.add_handler(CallbackQueryHandler(handle_symbol_selection, pattern=".*HZ.*"))
    application.add_handler(CallbackQueryHandler(handle_start_trading, pattern="start_trading"))
    application.add_handler(CallbackQueryHandler(handle_stop_trading, pattern="stop_trading"))

    application.run_polling()