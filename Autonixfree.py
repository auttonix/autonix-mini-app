import asyncio
import nest_asyncio
import websockets
import json
import html
import os
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, ContextTypes
)
from telegram.ext.filters import TEXT

nest_asyncio.apply()

WEBSOCKET_URL = "wss://ws.binaryws.com/websockets/v3?app_id=62714"

TELEGRAM_BOT_TOKEN = "7631264627:AAFXf02Z81OZZP9FEysfXVPK3TkWOjEcTU4"

LOG_FILE = "user_activityfree.xlsx"
if not os.path.exists(LOG_FILE) or not LOG_FILE.endswith(".xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "UserActivity"
    ws.append(["User", "Account Type", "Initial Balance", "Loss/Profit", "Total Profit", "Final Balance"])
    wb.save(LOG_FILE)

def log_user_activity(user_name, account_type, initial_balance, profit_loss, total_profit, final_balance):
    """
    Logs user trading activity into an Excel file.
    """
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb["UserActivity"]
        ws.append([user_name, account_type, initial_balance, profit_loss, total_profit, final_balance])
        wb.save(LOG_FILE)
    except Exception as e:
        print(f"Error logging activity: {e}")
        
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    disclaimer_text = (
        "⚠️ *Disclaimer: Trading Involves Risk*\n\n"
        "Trading binary options and other financial instruments on platforms like Deriv involves significant risk of loss and may not be suitable for all investors.\n\n"
        "📉 You could lose all your invested capital.\n"
        "📊 Ensure you fully understand the risks before proceeding.\n\n"
        "*Do you agree to the terms and wish to continue using the Autonix bot?*"
    )

    keyboard = [
        [InlineKeyboardButton("✅ I Agree", callback_data="agree_terms")],
        [InlineKeyboardButton("❌ I Disagree", callback_data="disagree_terms")],
        [InlineKeyboardButton("📞 Contact Customer Care", url="https://t.me/autonix_assistant_bot")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(disclaimer_text, parse_mode="Markdown", reply_markup=reply_markup)



async def handle_terms_agreement(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    user_data = context.user_data

    if query.data == "agree_terms":
        user_data["agreed_to_terms"] = True
        await query.edit_message_text(
            "✅ You have agreed to the trading terms.\n\n"
            "📈 Welcome to Autonix — Your Intelligent Trading Assistant 📈\n\n"
            "We're thrilled to have you onboard.\n\n"
            "To begin your journey with Autonix:\n"
            "✅ Ensure you have a valid *Deriv API Token\n"
            "✅ Configure your *stake amount*, *take profit, and stop loss\n"
            "✅ Select a market and initiate trading with precision\n\n"
            "💼 Designed for traders who value automation, accuracy, and control.\n\n"
            "Get started and take your trading to the next level.\n"
            "🔒 Your data is private. Your trades are yours.\n\n"
            "🧠 Powered by Autonix — Smart Trading Simplified."
            "You may now use the bot. Use /trade to get started."
        )
    else:
        user_data["agreed_to_terms"] = False
        await query.edit_message_text(
            "❌ You must agree to the disclaimer to use this bot.\n"
            "Restart anytime with /start."
        )



async def trade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.user_data.get("agreed_to_terms"):
        await update.message.reply_text(
            "⚠️ You must agree to the disclaimer before using this feature.\n"
            "Please use /start and accept the terms to proceed."
        )
        return

    keyboard = [
        [
            InlineKeyboardButton("📝 Create Deriv Account", url="https://track.deriv.com/_AmZyUaWfEMJB4VdSfJsOp2Nd7ZgqdRLk/1"),
            InlineKeyboardButton("🔐 Get API Token", url="https://app.deriv.com/account/api-token")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🔑 Please enter your Deriv API token to proceed.\n\n"
        "If you don't have one, use the buttons below to create an account or generate a token.",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    context.user_data["next_step"] = "api_token"


async def show_commands(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the /commands command and shows available commands."""
    commands_text = (
        "📜 *Available Commands:*\n\n"
        "🔹 /start - Introduction to the bot\n"
        "🔹 /trade - Begin trading authentication\n"
        "🔹 /commands - Show this list of commands\n"
    )
    await update.message.reply_text(commands_text, parse_mode="Markdown")

async def handle_user_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Processes sequential user input during trading setup."""

    user_input = update.message.text.strip()
    next_step = context.user_data.get("next_step")

    if not next_step:
        await update.message.reply_text(
            "⚠️ To begin trading, please use /start followed by the /trade command."
        )
        return

    # Step 1: API Token Authorization
    if next_step == "api_token":
        context.user_data["api_token"] = user_input

        try:
            async with websockets.connect(WEBSOCKET_URL) as websocket:
                await websocket.send(json.dumps({"authorize": user_input}))
                response = await websocket.recv()
                auth_data = json.loads(response)

                if "error" in auth_data:
                    error_msg = auth_data["error"].get("message", "Unknown error.")
                    await update.message.reply_text(
                        f"❌ *Authorization Failed:*\n_{error_msg}_",
                        parse_mode="Markdown"
                    )
                    return

                user_info = auth_data.get("authorize", {})
                context.user_data.update({
                    "user_name": user_info.get("fullname", "Trader"),
                    "account_balance": user_info.get("balance", 0),
                    
                })

                await update.message.reply_text(
                    f"✅ *API Token Verified Successfully!* 🔐\n\n"
                    f"👤 *User:* {context.user_data['user_name']}\n"
                    f"💼 *Account Type:* {context.user_data['account_type']}\n"
                    f"💰 *Balance:* {context.user_data['account_balance']:.2f} USD\n\n"
                    "Please enter your *stake amount* (e.g., `100`).",
                    parse_mode="Markdown"
                )
                context.user_data["next_step"] = "stake"

        except Exception as e:
            await update.message.reply_text(
                f"🚫 *Connection Error:*\n_Unable to connect to trading server. Please try again later._",
                parse_mode="Markdown"
            )
            return

    # Step 2: Stake Amount
    elif next_step == "stake":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("⚠️ Invalid input. Please enter a numeric value for the stake amount.")
            return

        context.user_data["stake"] = float(user_input)
        context.user_data["next_step"] = "take_profit"

        await update.message.reply_text(
            "✅ Stake amount saved.\nPlease enter your *take profit target* (e.g., `50`).",
            parse_mode="Markdown"
        )

    # Step 3: Take Profit Target
    elif next_step == "take_profit":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("⚠️ Invalid input. Please enter a numeric value for take profit.")
            return

        context.user_data["take_profit"] = float(user_input)
        context.user_data["next_step"] = "stop_loss"

        await update.message.reply_text(
            "✅ Take profit target set.\nNow enter your *stop loss limit* (e.g., `1000`).",
            parse_mode="Markdown"
        )

    # Step 4: Stop Loss Limit
    elif next_step == "stop_loss":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("⚠️ Invalid input. Please enter a numeric value for stop loss.")
            return

        context.user_data["stop_loss"] = float(user_input)
        context.user_data["next_step"] = None

        await update.message.reply_text(
            "🎯 All parameters configured successfully!\nPlease choose a trading symbol to proceed:",
            reply_markup=symbol_selection_keyboard()
        )


def symbol_selection_keyboard():
    """Returns a keyboard for selecting a trading symbol."""
    buttons = [
        [InlineKeyboardButton("Volatility 10 (1s)", callback_data="1HZ10V"),
         InlineKeyboardButton("Volatility 10", callback_data="R_10")],
        [InlineKeyboardButton("Volatility 25 (1s)", callback_data="1HZ25V"),
         InlineKeyboardButton("Volatility 25", callback_data="R_25")],
        [InlineKeyboardButton("Volatility 50 (1s)", callback_data="1HZ50V"),
         InlineKeyboardButton("Volatility 50", callback_data="R_50")],
        [InlineKeyboardButton("Volatility 75 (1s)", callback_data="1HZ75V"),
         InlineKeyboardButton("Volatility 75", callback_data="R_75")],
        [InlineKeyboardButton("Volatility 100 (1s)", callback_data="1HZ100V"),
         InlineKeyboardButton("Volatility 100", callback_data="R_100")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_symbol_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles symbol selection by the user."""
    query = update.callback_query
    await query.answer()
    symbol = query.data

    context.user_data["symbol"] = symbol
    await query.edit_message_text(
        text=f"✅ Symbol selected: {symbol}\n\n"
        "Now, use the buttons below to start or stop trading.",
        reply_markup=start_stop_trading_keyboard()
    )

def start_stop_trading_keyboard():
    """Returns a keyboard for starting or stopping trading."""
    buttons = [
        [InlineKeyboardButton("🚀 Start Trading", callback_data="start_trading")],
        [InlineKeyboardButton("🛑 Stop Trading", callback_data="stop_trading")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_trade_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles Start/Stop Trading actions."""
    query = update.callback_query
    await query.answer()
    
    action = query.data
    if action == "start_trading":
        await query.edit_message_text("🚀 Trading started! Executing trades now...")
        await execute_trades(update, context)
    elif action == "stop_trading":
        context.user_data["trading_active"] = False
        await query.edit_message_text("🛑 Trading stopped!")

async def execute_trades(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Executes trades using a modified Martingale strategy (stake × 4.4 on loss).
    Detects digit patterns in live tick data and places trades using Over/Under logic.
    Displays last tick digit in results and reverts stake after recovery win.
    """
    user_info = context.user_data
    context.user_data["trading_active"] = True

    # Extract user configuration
    symbol = user_info.get("symbol")
    stake = user_info.get("stake", 0)
    original_stake = stake
    take_profit = user_info.get("take_profit", 0)
    stop_loss = user_info.get("stop_loss", 0)
    api_token = user_info.get("api_token")
    user_name = user_info.get("user_name", "User")

    # Account info
    chat_id = update.effective_chat.id
    balance = user_info.get("account_balance", 0)
    initial_balance = balance
    total_profit = 0
    previous_was_loss = False

    await context.bot.send_message(chat_id, "🔍 Gathering market data to analyze patterns... Please wait ⏳")

    try:
        async with websockets.connect(WEBSOCKET_URL) as websocket:
            # Authorize user
            await websocket.send(json.dumps({"authorize": api_token}))
            while True:
                auth_data = json.loads(await websocket.recv())
                if auth_data.get("msg_type") == "authorize":
                    break
                elif "error" in auth_data:
                    await context.bot.send_message(chat_id, f"❌ Authorization failed: {auth_data['error']['message']}")
                    return

            # Subscribe to live ticks
            await websocket.send(json.dumps({"ticks": symbol, "subscribe": 1}))
            tick_window = []

            while context.user_data.get("trading_active", False):
                msg = json.loads(await websocket.recv())
                if msg.get("msg_type") != "tick":
                    continue

                # Collect tick digits
                tick = msg["tick"]
                last_digit = int(str(tick["quote"])[-1])
                tick_window.append(last_digit)

                if len(tick_window) > 10:
                    tick_window.pop(0)

                if len(tick_window) < 10:
                    continue

                # Analyze pattern
                over_count = sum(1 for d in tick_window if d > 1)
                under_count = sum(1 for d in tick_window if d < 8)

                if over_count > under_count:
                    contract_type, barrier = "DIGITOVER", "1"
                elif under_count > over_count:
                    contract_type, barrier = "DIGITUNDER", "8"
                else:
                    await context.bot.send_message(chat_id, "⚖️ No strong pattern detected. Skipping this round...")
                    await asyncio.sleep(2)
                    continue

                # Notify and prepare trade
                await context.bot.send_message(
                    chat_id,
                    f"📈 Signal Detected: {contract_type} {barrier} | Stake: ${stake:.2f}\n⏳ Placing trade..."
                )

                trade_request = {
                    "buy": 1,
                    "price": f"{stake:.2f}",
                    "parameters": {
                        "amount": f"{stake:.2f}",
                        "basis": "stake",
                        "contract_type": contract_type,
                        "currency": "USD",
                        "barrier": barrier,
                        "duration": 1,
                        "duration_unit": "t",
                        "symbol": symbol
                    }
                }

                await websocket.send(json.dumps(trade_request))

                # Wait for trade confirmation
                while True:
                    trade_data = json.loads(await websocket.recv())
                    if trade_data.get("msg_type") == "buy":
                        break

                if "error" in trade_data:
                    await context.bot.send_message(chat_id, f"❌ Trade failed: {trade_data['error']['message']}")
                    return

                if "buy" not in trade_data or "contract_id" not in trade_data["buy"]:
                    await context.bot.send_message(chat_id, "⚠️ Unexpected response. Trade not confirmed.")
                    return

                contract_id = trade_data["buy"]["contract_id"]

                await context.bot.send_message(
                    chat_id,
                    f"✅ Trade Executed: <b>{contract_type} {barrier}</b>\n"
                    f"💵 Stake: <b>${stake:.2f}</b>",
                    parse_mode="HTML"
                )

                # Monitor result
                while True:
                    await websocket.send(json.dumps({"proposal_open_contract": 1, "contract_id": contract_id}))
                    result_data = json.loads(await websocket.recv())

                    if result_data.get("msg_type") != "proposal_open_contract":
                        continue

                    if "error" in result_data:
                        await context.bot.send_message(chat_id, f"❌ Error checking trade result: {result_data['error']['message']}")
                        return

                    contract = result_data.get("proposal_open_contract", {})
                    if contract.get("is_sold"):
                        payout = contract.get("profit", 0)
                        total_profit += payout
                        balance += payout
                        context.user_data["account_balance"] = balance

                        is_win = payout >= 0
                        last_tick_digit = int(str(contract.get("exit_tick"))[-1]) if contract.get("exit_tick") else "N/A"

                        # Adjust stake after loss or recovery win
                        if previous_was_loss and is_win:
                            stake = original_stake
                            previous_was_loss = False
                        elif not is_win:
                            stake = round(stake * 4.4, 2)
                            previous_was_loss = True

                        # Result message
                        await context.bot.send_message(
                            chat_id,
                            f"{'🏆 Win!' if is_win else '❌ Loss!'}\n"
                            f"🎯 Last Tick Digit: <b>{last_tick_digit}</b>\n"
                            f"💰 Payout: <b>${payout:.2f}</b>\n"
                            f"📊 Total Profit: <b>${total_profit:.2f}</b>\n"
                            f"💼 Balance: <b>${balance:.2f}</b>",
                            parse_mode="HTML"
                        )

                        # Log trade
                        log_user_activity(user_name, user_info.get("account_type"), initial_balance, payout, total_profit, balance)

                        # Check for take profit or stop loss
                        if total_profit >= take_profit:
                            reply_markup = InlineKeyboardMarkup([
                                [InlineKeyboardButton("🌟 Get All Our Tools", url="https://t.me/autonix_bot")],
                                [InlineKeyboardButton("📊 Try Lite Bot", url="https://t.me/autonix_lite_bot")],
                                [InlineKeyboardButton("⚙️ Try Pro Bot", url="https://t.me/autonix_pro_bot")],
                                [InlineKeyboardButton("📢 Join Telegram Channel", url="https://t.me/autonix001")]
                            ])
                            await context.bot.send_message(
                                chat_id,
                                f"🎯 <b>Take Profit Target Achieved!</b>\n"
                                f"Well done, <i>{user_name}</i>! You secured profits and completed your session. 🎉",
                                parse_mode="HTML",
                                reply_markup=reply_markup
                            )
                            context.user_data["trading_active"] = False
                            return

                        elif total_profit <= -stop_loss:
                            await context.bot.send_message(chat_id, "🛑 Stop Loss hit. Ending trading session.")
                            context.user_data["trading_active"] = False
                            return

                        break

                await asyncio.sleep(3)

    except Exception as e:
        await context.bot.send_message(chat_id, f"❌ An unexpected error occurred: {str(e)}")
        context.user_data["trading_active"] = False


async def stop_trading(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Stops trading and resets trading state."""
    context.user_data["trading_active"] = False
    await update.message.reply_text("🛑 Trading has been stopped manually.")

def start_free_bot():
    """Start the bot."""
    print("Starting the free bot...")
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_terms_agreement, pattern="^(agree_terms|disagree_terms)$"))
    app.add_handler(CommandHandler("trade", trade))
    app.add_handler(CommandHandler("commands", show_commands))
    app.add_handler(CommandHandler("stop", stop_trading))
    app.add_handler(CallbackQueryHandler(handle_symbol_selection, pattern="^(R_|1HZ)"))
    app.add_handler(CallbackQueryHandler(handle_trade_action, pattern="^(start_trading|stop_trading)"))
    app.add_handler(MessageHandler(TEXT, handle_user_input))
    app.run_polling() 