import asyncio
import nest_asyncio
import websockets
import json
import html
import os
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, ContextTypes
)
from telegram.ext.filters import TEXT

nest_asyncio.apply()

WEBSOCKET_URL = "wss://ws.binaryws.com/websockets/v3?app_id=62714"

TELEGRAM_BOT_TOKEN = "7972065517:AAEg-rHfqfZzDF7p48rh22vlIWbyzUcc9AM"

TRADING_KEY = "autonixedgekey"

LOG_FILE = "user_activityedge.xlsx"
if not os.path.exists(LOG_FILE) or not LOG_FILE.endswith(".xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "UserActivity"
    ws.append(["User", "Account Type", "Initial Balance", "Loss/Profit", "Total Profit", "Final Balance"])
    wb.save(LOG_FILE)

def log_user_activity(user_name, account_type, initial_balance, profit_loss, total_profit, final_balance):
    """
    Logs user trading activity into an Excel file.
    """
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb["UserActivity"]
        ws.append([user_name, account_type, initial_balance, profit_loss, total_profit, final_balance])
        wb.save(LOG_FILE)
    except Exception as e:
        print(f"Error logging activity: {e}")
        
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    disclaimer_text = (
        "⚠️ *Disclaimer: Trading Involves Risk*\n\n"
        "Trading binary options and other financial instruments on platforms like Deriv involves significant risk of loss and may not be suitable for all investors.\n\n"
        "📉 You could lose all your invested capital.\n"
        "📊 Ensure you fully understand the risks before proceeding.\n\n"
        "*Do you agree to the terms and wish to continue using the Autonix bot?*"
    )

    keyboard = [
        [InlineKeyboardButton("✅ I Agree", callback_data="agree_terms")],
        [InlineKeyboardButton("❌ I Disagree", callback_data="disagree_terms")],
        [InlineKeyboardButton("📞 Contact Customer Care", url="https://t.me/autonix_assistant_bot")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(disclaimer_text, parse_mode="Markdown", reply_markup=reply_markup)

async def handle_terms_agreement(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    user_data = context.user_data

    if query.data == "agree_terms":
        user_data["agreed_to_terms"] = True
        await query.edit_message_text(
            "✅ You have agreed to the trading terms.\n\n"
            "📈 Welcome to Autonixedge — Your Intelligent Trading Assistant 📈\n\n"
            "To begin your journey with Autonix:\n"
            "✅ Ensure you have a valid *Deriv API Token\n"
            "✅ Configure your *stake amount*, *take profit, and stop loss\n"
            "✅ Select a market and initiate trading with precision\n\n"
            "💼 Designed for traders who value automation, accuracy, and control.\n\n"
            "Get started and take your trading to the next level.\n"
            "🔒 Your data is private. Your trades are yours.\n\n"
            "🧠 Powered by Autonix — Smart Trading Simplified."
            "You may now use the bot. Use /trade to get started."
        )
    else:
        user_data["agreed_to_terms"] = False
        await query.edit_message_text(
            "❌ You must agree to the disclaimer to use this bot.\n"
            "Restart anytime with /start."
        )

async def trade(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.user_data.get("agreed_to_terms"):
        await update.message.reply_text(
            "⚠️ You must agree to the disclaimer before using this feature.\n"
            "Please use /start and accept the terms to proceed."
        )
        return

    keyboard = [
        [
            InlineKeyboardButton("📝 Create Deriv Account", url="https://track.deriv.com/_AmZyUaWfEMJB4VdSfJsOp2Nd7ZgqdRLk/1"),
            InlineKeyboardButton("🔐 Get API Token", url="https://app.deriv.com/account/api-token")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(
        "🔑 Please enter your Deriv API token to proceed.\n\n"
        "If you don't have one, use the buttons below to create an account or generate a token.",
        parse_mode="Markdown",
        reply_markup=reply_markup
    )

    context.user_data["next_step"] = "api_token"

async def show_commands(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the /commands command and shows available commands."""
    commands_text = (
        "📜 *Available Commands:*\n\n"
        "🔹 /start - Introduction to the bot\n"
        "🔹 /trade - Begin trading authentication\n"
        "🔹 /commands - Show this list of commands\n"
    )
    await update.message.reply_text(commands_text, parse_mode="Markdown")

async def handle_user_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_input = update.message.text
    next_step = context.user_data.get("next_step")
    
    if not next_step:
        await update.message.reply_text("Please start the bot using /start and then click /trade.")
        return
    
    if next_step == "api_token":
        api_token = user_input
        context.user_data["api_token"] = api_token
        async with websockets.connect(WEBSOCKET_URL) as websocket:
            await websocket.send(json.dumps({"authorize": api_token}))
            response = await websocket.recv()
            auth_data = json.loads(response)
            
            if "error" in auth_data:
                await update.message.reply_text(f"Authorization failed: {auth_data['error']['message']}")
                return
            
            user_name = auth_data.get("authorize", {}).get("fullname", "User")
            account_balance = auth_data.get("authorize", {}).get("balance", 0)
            context.user_data["user_name"] = user_name
            context.user_data["account_balance"] = account_balance
            
            await update.message.reply_text(
                f"✅ API Token Authenticated Successfully! 🎉\n\n"
                f"👤 User: {user_name}\n"
                f"💰 Balance: {account_balance} USD\n\n"
                "🔑 Now, enter your trading key to continue.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("I don't have a trading key", url="https://t.me/Autonixedgekey_bot")]])
            )
        context.user_data["next_step"] = "trading_key"

    elif next_step == "trading_key":
        if user_input == TRADING_KEY:
            await update.message.reply_text(
                "✅ Trading key authenticated successfully! 🎉\n\n"
                "📊 Select the trading symbol you want to trade:",
                reply_markup=symbol_selection_keyboard()
            )
            context.user_data["next_step"] = "symbol"
        else:
            await update.message.reply_text("Invalid key. Please enter a correct trading key.")

    elif next_step == "stake":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the stake amount.")
            return
        context.user_data["stake"] = float(user_input)
        await update.message.reply_text("Stake saved! Now, enter your take 💰profit target✅.")
        context.user_data["next_step"] = "take_profit"

    elif next_step == "take_profit":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the take profit target.")
            return
        context.user_data["take_profit"] = float(user_input)
        await update.message.reply_text("Take profit saved! Now, enter your stop loss target.")
        context.user_data["next_step"] = "stop_loss"

    elif next_step == "stop_loss":
        if not user_input.replace('.', '', 1).isdigit():
            await update.message.reply_text("Please enter a valid number for the stop loss target.")
            return
        context.user_data["stop_loss"] = float(user_input)
        await update.message.reply_text(
            "Setup complete! Use the buttons below to start or stop trading.",
            reply_markup=start_stop_trading_keyboard(),
        )
        context.user_data["next_step"] = None

def symbol_selection_keyboard():
    """Returns a keyboard for selecting a trading symbol."""
    buttons = [
        [InlineKeyboardButton("Volatility 10 (1s)", callback_data="1HZ10V"),
         InlineKeyboardButton("Volatility 10", callback_data="R_10")],
        [InlineKeyboardButton("Volatility 25 (1s)", callback_data="1HZ25V"),
         InlineKeyboardButton("Volatility 25", callback_data="R_25")],
        [InlineKeyboardButton("Volatility 50 (1s)", callback_data="1HZ50V"),
         InlineKeyboardButton("Volatility 50", callback_data="R_50")],
        [InlineKeyboardButton("Volatility 75 (1s)", callback_data="1HZ75V"),
         InlineKeyboardButton("Volatility 75", callback_data="R_75")],
        [InlineKeyboardButton("Volatility 100 (1s)", callback_data="1HZ100V"),
         InlineKeyboardButton("Volatility 100", callback_data="R_100")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_symbol_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    symbol = query.data

    context.user_data["symbol"] = symbol
    await query.edit_message_text(
        text=f"Symbol selected: {symbol}\n\n"
        "Please enter your stake amount (e.g., 10)."
    )
    context.user_data["next_step"] = "stake"

def start_stop_trading_keyboard():
    buttons = [
        [InlineKeyboardButton("🚀 Start Trading", callback_data="start_trading")],
        [InlineKeyboardButton("🛑 Stop Trading", callback_data="stop_trading")],
    ]
    return InlineKeyboardMarkup(buttons)

async def handle_trade_action(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles Start/Stop Trading actions."""
    query = update.callback_query
    await query.answer()
    
    action = query.data
    if action == "start_trading":
        await query.edit_message_text("🚀 Trading started! Executing trades now...")
        await execute_trades(update, context)
    elif action == "stop_trading":
        context.user_data["trading_active"] = False
        await query.edit_message_text("🛑 Trading stopped!")

async def execute_trades(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_info = context.user_data
    context.user_data["trading_active"] = True

    # Extract user settings
    symbol = user_info.get("symbol")
    original_stake = user_info.get("stake", 0)
    stake = original_stake
    take_profit = user_info.get("take_profit", 0)
    stop_loss = user_info.get("stop_loss", 0)
    api_token = user_info.get("api_token")
    user_name = user_info.get("user_name", "User")
    chat_id = update.effective_chat.id

    # Initialize trading session variables
    balance = user_info.get("account_balance", 0)
    initial_balance = balance
    total_profit = 0
    accumulated_loss = 0
    in_recovery_mode = False
    recovery_payout = 0

    await context.bot.send_message(chat_id, "🔍 Starting market analysis for <b>DIGITDIFF</b> strategy...", parse_mode="HTML")

    try:
        async with websockets.connect(WEBSOCKET_URL) as websocket:
            # Authorize
            await websocket.send(json.dumps({"authorize": api_token}))
            while True:
                auth_data = json.loads(await websocket.recv())
                if auth_data.get("msg_type") == "authorize":
                    break
                elif "error" in auth_data:
                    await context.bot.send_message(chat_id, f"❌ Authorization failed: {auth_data['error']['message']}")
                    return

            # Subscribe to ticks
            await websocket.send(json.dumps({"ticks": symbol, "subscribe": 1}))
            tick_window = []

            while context.user_data.get("trading_active", False):
                msg = json.loads(await websocket.recv())
                if msg.get("msg_type") != "tick":
                    continue

                tick = msg["tick"]
                last_digit = int(str(tick["quote"])[-1])
                tick_window.append(last_digit)

                if len(tick_window) > 20:
                    tick_window.pop(0)
                if len(tick_window) < 10:
                    continue

                # Select strategy
                if not in_recovery_mode:
                    digit_counts = {i: tick_window.count(i) for i in range(10)}
                    barrier = str(max(digit_counts, key=digit_counts.get))
                    contract_type = "DIGITDIFF"
                    stake = original_stake

                    await context.bot.send_message(
                        chat_id,
                        f"📈 Analyzed last 20 ticks\n"
                        f"🔥 Now differing digit: <b>{barrier}</b>\n"
                        f"🎯 Strategy: <b>{contract_type}</b> (≠ {barrier})",
                        parse_mode="HTML"
                    )
                else:
                    # Recovery mode with digit trend analysis
                    digit_freq = {i: tick_window.count(i) for i in range(10)}
                    over_score = sum(freq * digit for digit, freq in digit_freq.items() if digit > 1)
                    under_score = sum(freq * (9 - digit) for digit, freq in digit_freq.items() if digit < 8)

                    if over_score > under_score:
                        contract_type = "DIGITOVER"
                        barrier = "1"
                        trend_direction = "📈 Higher digits dominating"
                    else:
                        contract_type = "DIGITUNDER"
                        barrier = "8"
                        trend_direction = "📉 Lower digits dominating"

                    await context.bot.send_message(
                        chat_id,
                        f"♻️ Recovery Mode Analysis:\n"
                        f"{trend_direction}\n"
                        f"🧠 Recovery Strategy: <b>{contract_type} {barrier}</b>",
                        parse_mode="HTML"
                    )

                # Execute trade
                await context.bot.send_message(
                    chat_id,
                    f"📡 Trade Signal: {contract_type} {barrier} | Stake: ${stake:.2f}\n⏳ Executing trade..."
                )

                trade_request = {
                    "buy": 1,
                    "price": f"{stake:.2f}",
                    "parameters": {
                        "amount": f"{stake:.2f}",
                        "basis": "stake",
                        "contract_type": contract_type,
                        "currency": "USD",
                        "barrier": barrier,
                        "duration": 1,
                        "duration_unit": "t",
                        "symbol": symbol
                    }
                }

                await websocket.send(json.dumps(trade_request))
                while True:
                    trade_data = json.loads(await websocket.recv())
                    if trade_data.get("msg_type") == "buy":
                        break

                if "error" in trade_data:
                    await context.bot.send_message(chat_id, f"❌ Trade failed: {trade_data['error']['message']}")
                    return

                contract_id = trade_data["buy"]["contract_id"]

                await context.bot.send_message(
                    chat_id,
                    f"✅ Trade Executed: <b>{contract_type} {barrier}</b>\n"
                    f"💵 Stake: <b>${stake:.2f}</b>",
                    parse_mode="HTML"
                )

                # Monitor result
                while True:
                    await websocket.send(json.dumps({"proposal_open_contract": 1, "contract_id": contract_id}))
                    result_data = json.loads(await websocket.recv())

                    if result_data.get("msg_type") != "proposal_open_contract":
                        continue
                    if "error" in result_data:
                        await context.bot.send_message(chat_id, f"❌ Error checking result: {result_data['error']['message']}")
                        return

                    contract = result_data.get("proposal_open_contract", {})
                    if contract.get("is_sold"):
                        payout = contract.get("profit", 0)
                        is_win = payout >= 0
                        last_tick_digit = int(str(contract.get("exit_tick"))[-1]) if contract.get("exit_tick") else "N/A"

                        total_profit += payout
                        balance += payout
                        context.user_data["account_balance"] = balance

                        await context.bot.send_message(
                            chat_id,
                            f"{'🏆 Win!' if is_win else '❌ Loss!'}\n"
                            f"🎯 Last Tick Digit: <b>{last_tick_digit}</b>\n"
                            f"💰 Payout: <b>${payout:.2f}</b>\n"
                            f"📊 Total Profit: <b>${total_profit:.2f}</b>\n"
                            f"💼 Balance: <b>${balance:.2f}</b>",
                            parse_mode="HTML"
                        )

                        log_user_activity(user_name, user_info.get("account_type"), initial_balance, payout, total_profit, balance)

                        # Handle recovery logic
                        if not in_recovery_mode and not is_win:
                            accumulated_loss = abs(payout)
                            stake = round(accumulated_loss, 2)
                            in_recovery_mode = True
                            recovery_payout = 0
                            await context.bot.send_message(chat_id, f"⚠️ Loss occurred. Entering recovery mode.\n🔁 Total to recover: ${accumulated_loss:.2f}")

                        elif in_recovery_mode:
                            if is_win:
                                recovery_payout += payout
                                remaining = accumulated_loss - recovery_payout
                                if remaining <= 0:
                                    in_recovery_mode = False
                                    accumulated_loss = 0
                                    recovery_payout = 0
                                    stake = original_stake
                                    await context.bot.send_message(chat_id, "✅ Full recovery complete! Returning to main strategy.")
                                else:
                                    stake += payout
                                    await context.bot.send_message(chat_id, f"💵 Partial recovery. Remaining to recover: ${remaining:.2f}")
                            else:
                                accumulated_loss += abs(payout)
                                stake = round(accumulated_loss, 2)
                                recovery_payout = 0
                                await context.bot.send_message(chat_id, f"❌ Recovery loss! New recovery target: ${accumulated_loss:.2f}")

                        # Check stop conditions
                        if total_profit >= take_profit:
                            reply_markup = InlineKeyboardMarkup([
                                [InlineKeyboardButton("🌟 Get All Our Tools", url="https://t.me/autonix_bot")],
                                [InlineKeyboardButton("🌟 Try our free Bot", url="https://t.me/autonix_free_bot")],
                                [InlineKeyboardButton("📊 Try Lite Bot", url="https://t.me/autonix_lite_bot")],
                                [InlineKeyboardButton("⚙️ Try Pro Bot", url="https://t.me/autonix_pro_bot")],
                                [InlineKeyboardButton("📢 Join Telegram Channel", url="https://t.me/autonix001")]
                            ])
                            await context.bot.send_message(
                                chat_id,
                                f"🎯 <b>Take Profit Target Achieved!</b>\n"
                                f"Well done, <i>{user_name}</i>! You secured profits and completed your session. 🎉",
                                parse_mode="HTML",
                                reply_markup=reply_markup
                            )
                            context.user_data["trading_active"] = False
                            return

                        elif total_profit <= -stop_loss:
                            context.user_data["trading_active"] = False
                            await context.bot.send_message(chat_id, "🛑 Stop Loss triggered. Session ended.")
                            return

                        break

                await asyncio.sleep(2)

    except Exception as e:
        context.user_data["trading_active"] = False
        await context.bot.send_message(chat_id, f"❌ Error occurred: {str(e)}")

async def stop_trading(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Stops trading and resets trading state."""
    context.user_data["trading_active"] = False
    await update.message.reply_text("🛑 Trading has been stopped manually.")

def start_AutonixEdge_bot():
    """Start the bot."""
    print("Starting the Autonixedge bot...")
    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(handle_terms_agreement, pattern="^(agree_terms|disagree_terms)$"))
    app.add_handler(CommandHandler("trade", trade))
    app.add_handler(CommandHandler("commands", show_commands))
    app.add_handler(CommandHandler("stop", stop_trading))
    app.add_handler(CallbackQueryHandler(handle_symbol_selection, pattern="^(R_|1HZ)"))
    app.add_handler(CallbackQueryHandler(handle_trade_action, pattern="^(start_trading|stop_trading)"))
    app.add_handler(MessageHandler(TEXT, handle_user_input))
    app.run_polling() 